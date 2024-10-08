# Importar Las Librerias
import PyPDF2
import os
import shutil
import re
import time
import docx2pdf
from dateutil.relativedelta import relativedelta
from datetime import datetime, date
from docxtpl import DocxTemplate
from pathlib import Path
import openpyxl
import sys
import mimetypes

# Carpeta donde se encuentran los PDF...
# carpeta_pdf = 'C:/CotizacionDescarga/'
carpeta_pdf = os.path.join(os.environ['USERPROFILE'], 'Downloads')
# Carpeta donde se moverán los PDF procesados
carpeta_procesados = 'C:/CotizacionProcesada/'
# Plantilla de Axeso
plantilla_axeso = 'C:/ExtractPDF/Axeso.docx'

class CondicionesAxeso:
    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.wb = openpyxl.load_workbook(archivo_excel)
        self.sheet = self.wb['Condiciones']

    def carga_datos(self):
        self.monto_inicial = self.sheet.cell(row=2, column=1).value
        self.numero_cuotas = self.sheet.cell(row=2, column=2).value
        self.meses_a_sumar = self.sheet.cell(row=2, column=3).value
        self.nrocotiza = self.sheet.cell(row=2, column=4).value + 1

    def pago_mensual(self,monto):
        # Define el límite superior
        limite_superior = 9999999

        # Define las celdas que contienen los rangos y factores
        rangos = []
        factores = []

        # Lee los valores de las celdas
        for row in range(2, 18):  # desde la fila 2 hasta la 17
            cell = self.sheet.cell(row=row, column=9)  # columna I es la número 9
            rangos.append(cell.value)

        for row in range(2, 18):  # desde la fila 2 hasta la 17
            cell = self.sheet.cell(row=row, column=11)  # columna K es la número 11
            factores.append(cell.value)

        # Implementa la lógica para comparar el monto y obtener el factor correspondiente
        for i, rango in enumerate(rangos):
            if monto <= rango:
                factor = factores[i]
                calculo = monto * factor
                break
            elif monto > limite_superior:
                factor = factores[-1]  # toma el factor del límite superior
                calculo = monto * factor
                break
        else:
            calculo = 0  # Si no se encuentra un rango que coincida, asigna None
        return calculo 

    def incrementa_nrocotiza(self):
        self.sheet.cell(row=2, column=4).value += 1

    def guarda_datos(self):
        self.wb.save(self.archivo_excel)

def extract_tipo_documento(pdf_file_path: Path) -> str:
    try:
        pdf = PyPDF2.PdfReader(str(pdf_file_path))
        page = pdf.pages[1]
        text = page.extract_text()
        #print('text:',text)
        # Creamos un patrón regex para buscar los valores
        pattern = re.compile(r'\b(QUOTE NO:|QUOTE NO.|POLICY NO:)\s*(\w+)\b', re.IGNORECASE)
        # Buscamos el patrón en el texto
        match = pattern.search(text)
        if match:
           if match.group(1) in ('QUOTE NO:','QUOTE NO.'):
              return match.group(1)
           else:
              return None
        else:
           return None
    except PyPDF2.PdfReadError as e:
        print(f"Error leyendo el archivo PDF: {e}")
        return None

      
def extract_client_info(pdf_file_path: Path) -> str:
    #Extrae información del cliente de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        insured_index = text.find('INSURED:')
        if insured_index != -1:
           insured_data = text[insured_index + 8:]
           effective_date_index = insured_data.find('EFFECTIVE DATE:')
           if effective_date_index != -1:
              insured_data = insured_data[:effective_date_index]
              # Remove newline characters
              insured_data = insured_data.replace('\n', ' ')
              # Remove double spaces
              insured_data = re.sub(' +', ' ', insured_data)
              return insured_data.strip()
    return None

def extract_direccion_info(pdf_file_path: Path) -> str:
    #Extrae información de la Direccion de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        insured_index = text.find('AGENCY AND MAILING ADDRESS 2')
        if insured_index != -1:
            insured_data = text[insured_index + 28:]
            singular_insurance_index = insured_data.find('SINGULAR INSURANCE AGENCY,')
            if singular_insurance_index != -1:
                insured_data = insured_data[:singular_insurance_index]
                # Remove newline characters
                insured_data = insured_data.replace('\n', ' ')
                # Remove double spaces
                insured_data = re.sub(' +', ' ', insured_data)
                return insured_data.strip()
    return None 

def extract_direccion_info_2(pdf_file_path: Path) -> str:
    #Extrae información de la Direccion de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        insured_index = text.find('AGENCY AND MAILING A DDRESS  2')
        if insured_index != -1:
            insured_data = text[insured_index + 31:]
            singular_insurance_index = insured_data.find('SINGULAR INSURANCE A GENCY,')
            if singular_insurance_index != -1:
                insured_data = insured_data[:singular_insurance_index]
                # Remove newline characters
                insured_data = insured_data.replace('\n', ' ')
                # Remove double spaces
                insured_data = re.sub(' +', ' ', insured_data)
                return insured_data.strip()
    return None  

def extract_poliza_info(pdf_file_path: Path) -> str:
    #Extrae información de la Direccion de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        # Creamos un patrón regex para buscar los valores
        pattern = re.compile(r'QUOTE NO:\s*(.*)')
        # Buscamos el patrón en el texto
        match = pattern.search(text)
        if match:
            # Remove newline characters
            insured_data = match.group(1)
            insured_data = insured_data.replace('\n', ' ')
            # Remove double spaces
            insured_data = re.sub(' +', ' ', insured_data)
            cadena = insured_data.strip()
            partes = cadena.split("-")
            partes = [parte.lstrip("0") or "0" for parte in partes]
            numero_poliza = "-".join(partes)
            numero_poliza = numero_poliza[:-2]
            return numero_poliza
    return '00-00-00'

def extract_prima_info(pdf_file_path: Path) -> float:
    try:
        pdf = PyPDF2.PdfReader(str(pdf_file_path))
        for page in pdf.pages:
            text = page.extract_text()
            pattern = r'(?:POLICY PREMIUM|GENERAL LIABILITY PREMIUM|ESTIMATED POLICY PREMIUM)\s*\$\s*([\d,\.]+)'
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value = match.group(1).replace(',', '')  # Removemos las comas
                try:
                    value = float(value)  # Intentamos convertir a float
                    return value
                except ValueError:
                    print(f"No se pudo convertir el valor '{value}' a float")
                    return None
    except PyPDF2.PdfReadError as e:
        print(f"Error leyendo el archivo PDF: {e}")
        return None
    
def extract_desde_info(pdf_file_path: Path) -> str:
    #Extrae información de la Fecha Desde de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        insured_index = text.find('FROM')
        if insured_index != -1:
            insured_data = text[insured_index + 4:]
            to_index = insured_data.find('TO')
            if to_index != -1:
                insured_data = insured_data[:to_index]
                # Remove newline characters
                insured_data = insured_data.replace('\n', ' ')
                # Remove double spaces
                insured_data = re.sub(' +', ' ', insured_data)
                return insured_data.strip()
    return None
        
def extract_fechapago1_info(pdf_file_path: Path) -> str:
    #Extrae información de la Fecha Desde de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    for page in pdf.pages:
        text = page.extract_text()
        insured_index = text.find('FROM')
        if insured_index != -1:
            insured_data = text[insured_index + 4:]
            to_index = insured_data.find('TO')
            if to_index != -1:
                insured_data = insured_data[:to_index]
                # Remove newline characters
                insured_data = insured_data.replace('\n', ' ')
                # Remove double spaces
                insured_data = re.sub(' +', ' ', insured_data)
                return insured_data.strip()
    return None
    
def extract_pdf_info(pdf_file_path: Path) -> dict:
    prima_info = extract_prima_info(pdf_file_path)
    # Crear una instancia de la clase
    condiciones_axeso = CondicionesAxeso('C:/ExtractPDF/CondicionesAxeso.xlsx')
    # Cargar datos
    condiciones_axeso.carga_datos()
    # Incrementar nrocotiza
    condiciones_axeso.incrementa_nrocotiza()
    # Buscar Cargo del Financiamiento
    deposito = prima_info * condiciones_axeso.monto_inicial
    ncuotas = condiciones_axeso.numero_cuotas
    cantfin = prima_info - deposito
    mtocuota = round(condiciones_axeso.pago_mensual(cantfin),2)
    #cargo = condiciones_axeso.cargo_financiamiento(cantfin)
    cargo = mtocuota * ncuotas - cantfin
    # Guardar datos
    condiciones_axeso.guarda_datos()
    #Extrae información de un archivo PDF
    client_info = extract_client_info(pdf_file_path)
    direccion_info = extract_direccion_info(pdf_file_path)
    if direccion_info is None:
       direccion_info = extract_direccion_info_2(pdf_file_path)

    poliza_info = extract_poliza_info(pdf_file_path)
    prima_info = extract_prima_info(pdf_file_path)
    desde_info = extract_desde_info(pdf_file_path)
    fechapago1 = date.today().strftime("%m/%d/%Y")
    fecha_pago1_date = datetime.strptime(fechapago1, '%m/%d/%Y')
    fechapago2_date = fecha_pago1_date + relativedelta(months=ncuotas)
    fechapago2 = fechapago2_date.strftime('%m/%d/%Y')
    totalpag = cantfin + cargo
    mtoventa = prima_info + cargo
    #print('client_info: ',client_info)
    #print('direccion_info: ', direccion_info)
    #print('poliza_info: ',poliza_info)
    #print('prima_info: ',prima_info)
    #print('desde_info',desde_info)
    if client_info and direccion_info and prima_info and desde_info:
        return {'cliente': client_info, 
                'direccion':direccion_info,
                'poliza':poliza_info,
                'prima': "{0:,.2f}".format(prima_info),
                'desde': desde_info,
                'fechapago1': fechapago1,
                'fechapago2': fechapago2,
                'deposito': "{0:,.2f}".format(deposito),
                'cuotas': ncuotas,
                'cantfin': "{0:,.2f}".format(cantfin),
                'nrocotiza': condiciones_axeso.nrocotiza,
                'cargo': "{0:,.2f}".format(cargo),
                'totalpag': "{0:,.2f}".format(totalpag),
                'mtocuota': "{0:,.2f}".format(mtocuota),
                'mtoventa': "{0:,.2f}".format(mtoventa)
                 }
 
def generate_docx(pdf_info: dict, nbredocpdf: str) -> None:
    #Genera un archivo DOCX a partir de la información extraída del PDF
    template = DocxTemplate(plantilla_axeso)
    datos = {
        'direccion': pdf_info['direccion'],
        'cliente': pdf_info['cliente'],
        'prima': pdf_info['prima'],
        'dep': pdf_info['deposito'],
        'cuotas': pdf_info['cuotas'],
        'fechapago1': pdf_info['fechapago1'],
        'fechapago2': pdf_info['fechapago2'],
        'desde': pdf_info['desde'],
        'poliza': pdf_info['poliza'],
        'cantfin': pdf_info['cantfin'],
        'nrocotiza': pdf_info['nrocotiza'],
        'cargo': pdf_info['cargo'],
        'totalpag': pdf_info['totalpag'],
        'mtocuota': pdf_info['mtocuota'],
        'mtoventa': pdf_info['mtoventa'],
    }

    template.render(datos)
    nombre_archivo = 'Axeso '+str(datos['nrocotiza'])+'.docx'
    #nombre_archivo_pdf = 'Axeso '+str(datos['nrocotiza'])+'.pdf'
    nombre_archivo_pdf = 'Axeso '+nbredocpdf

    ruta_completa = os.path.join(carpeta_procesados, nombre_archivo)
    ruta_completa_pdf = os.path.join(carpeta_procesados, nombre_archivo_pdf)

    # Si el documento existe en procesado. Se debe eliminar antes.
    if os.path.exists(ruta_completa_pdf):
       os.remove(os.path.join(carpeta_procesados, nombre_archivo_pdf))
    
    template.save(ruta_completa)
    
    # Convertir el documento a PDF
    docx2pdf.convert(ruta_completa, ruta_completa_pdf)

    # Elimina el archivo Word
    path_docx = os.path.join(carpeta_procesados, nombre_archivo)
    os.remove(path_docx)
    
def process_pdf_file(pdf_file_path: Path) -> None:
    # Procesa un archivo PDF y genera un archivo DOCX
    pdf_info = extract_pdf_info(pdf_file_path)
    if pdf_info:
        file_name = os.path.basename(pdf_file_path)
        file_path = os.path.join(carpeta_procesados, file_name)
        generate_docx(pdf_info,file_name)
        #print('name,path:',file_name,file_path)
        if os.path.exists(file_path):
            os.remove(os.path.join(carpeta_procesados, file_name))

        shutil.move(os.path.join(carpeta_pdf, pdf_file_path), carpeta_procesados)

def is_file_locked(file_path):
    try:
        #os.open(file_path, os.O_EXCL | os.O_RDWR)
        os.open(file_path, os.O_RDONLY)
    except OSError:
        return True
    return False        
        

def main() -> None:
    for file in os.listdir(carpeta_pdf):
        if file.endswith('.pdf'):
            pdf_file_path = Path(os.path.join(carpeta_pdf, file))
            #if not is_file_locked(pdf_file_path):
            mime_type, _ = mimetypes.guess_type(str(pdf_file_path))
            if mime_type == 'application/pdf':
               fecha_creacion = datetime.fromtimestamp(os.path.getctime(pdf_file_path)).date()
               #print('fecha creacion: ',fecha_creacion,date.today(),pdf_file_path)
               if fecha_creacion == date.today():
                  try:
                        tipo_doc = extract_tipo_documento(pdf_file_path)
                        if tipo_doc:
                            # Procesar Archivo PDF
                            process_pdf_file(pdf_file_path)
                  except Exception as e:
                         print(f"Error al leer el archivo PDF {pdf_file_path} {e}")
            else:
                print(f"El archivo {pdf_file_path} no es un PDF válido")          

if __name__ == "__main__":
    main()


            