# Importar Las Librerias
import PyPDF2
import os
import shutil
import re
import time
import docx2pdf
from dateutil.relativedelta import relativedelta
from datetime import datetime, date
from docxtpl import DocxTemplate
from pathlib import Path
import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Carga la hoja de Excel
wb = openpyxl.load_workbook('CondicionesAxeso.xlsx')

# Selecciona la hoja de trabajo
sheet = wb['Condiciones']

# Lee los datos
monto_inicial = sheet.cell(row=2, column=1).value
print('monto inicial:',monto_inicial)
numero_cuotas = sheet.cell(row=2, column=2).value
meses_a_sumar = sheet.cell(row=2, column=3).value

# Carpeta donde se encuentran los PDF
carpeta_pdf = 'C:/CotizacionDescarga/'
# Carpeta donde se moverán los PDF procesados
carpeta_procesados = 'C:/CotizacionProcesada/'
# Plantilla de Axeso
plantilla_axeso = 'Axeso.docx'

#meses_a_sumar = 9
#monto_inicial = 0.25
#numero_cuotas = 9

      
def extract_client_info(pdf_file_path: Path) -> str:
    #Extrae información del cliente de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[2]
    text = page.extract_text()
    insured_index = text.find('INSURED:')
    if insured_index != -1:
       insured_data = text[insured_index + 8:]
       effective_date_index = insured_data.find('EFFECTIVE DATE:')
       if effective_date_index != -1:
          insured_data = insured_data[:effective_date_index]
          # Remove newline characters
          insured_data = insured_data.replace('\n', ' ')
          # Remove double spaces
          insured_data = re.sub(' +', ' ', insured_data)
          return insured_data.strip()


def extract_direccion_info(pdf_file_path: Path) -> str:
    #Extrae información de la Direccion de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[1]
    text = page.extract_text()
    insured_index = text.find('AGENCY AND MAILING ADDRESS 2')
    if insured_index != -1:
        insured_data = text[insured_index + 28:]
        singular_insurance_index = insured_data.find('SINGULAR INSURANCE AGENCY,')
        if singular_insurance_index != -1:
            insured_data = insured_data[:singular_insurance_index]
        # Remove newline characters
        insured_data = insured_data.replace('\n', ' ')
        # Remove double spaces
        insured_data = re.sub(' +', ' ', insured_data)
        return insured_data.strip()

def extract_poliza_info(pdf_file_path: Path) -> str:
    #Extrae información de la Direccion de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[1]
    text = page.extract_text()
    insured_index = text.find('QUOTE NO:')
    if insured_index != -1:
        insured_data = text[insured_index + 9:]
        renewal_index = insured_data.find('RENEWAL')
        if renewal_index != -1:
            insured_data = insured_data[:renewal_index]
        # Remove newline characters
        insured_data = insured_data.replace('\n', ' ')
        # Remove double spaces
        insured_data = re.sub(' +', ' ', insured_data)
        cadena = insured_data.strip()
        partes = cadena.split("-")
        partes = [parte.lstrip("0") or "0" for parte in partes]
        numero_poliza = "-".join(partes)
        numero_poliza = numero_poliza[:-2]
        return numero_poliza

def extract_prima_info(pdf_file_path: Path) -> float:
    #Extrae información de prima de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[2]
    text = page.extract_text()
    prima_index = text.find('ESTIMATED GENERAL LIABILITY PREMIUM')
    if prima_index != -1:
        prima_data = text[prima_index + 36:]
        monto_index = prima_data.find('FORMS AND ENDORSEMENTS')
        if monto_index != -1:
            prima_data = prima_data[:monto_index]
        prima = float(prima_data.replace('\n', ' ').replace('$', '').strip())
        print('valor prima: ',prima)
        return prima
    
def extract_desde_info(pdf_file_path: Path) -> str:
    #Extrae información de la Fecha Desde de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[1]
    text = page.extract_text()
    insured_index = text.find('POLICY PERIOD: FROM')
    if insured_index != -1:
        insured_data = text[insured_index + 19:]
        to_index = insured_data.find('TO')
        if to_index != -1:
            insured_data = insured_data[:to_index]
        # Remove newline characters
        insured_data = insured_data.replace('\n', ' ')
        # Remove double spaces
        insured_data = re.sub(' +', ' ', insured_data)
        return insured_data.strip()
        
def extract_fechapago1_info(pdf_file_path: Path) -> str:
    #Extrae información de la Fecha Desde de un archivo PDF
    pdf = PyPDF2.PdfReader(str(pdf_file_path))
    page = pdf.pages[1]
    text = page.extract_text()
    insured_index = text.find('POLICY PERIOD: FROM')
    if insured_index != -1:
        insured_data = text[insured_index + 19:]
        to_index = insured_data.find('TO')
        if to_index != -1:
            insured_data = insured_data[:to_index]
        # Remove newline characters
        insured_data = insured_data.replace('\n', ' ')
        # Remove double spaces
        insured_data = re.sub(' +', ' ', insured_data)
        return insured_data.strip()
    
def extract_pdf_info(pdf_file_path: Path) -> dict:
    #Extrae información de un archivo PDF
    client_info = extract_client_info(pdf_file_path)
    direccion_info = extract_direccion_info(pdf_file_path)
    poliza_info = extract_poliza_info(pdf_file_path)
    prima_info = extract_prima_info(pdf_file_path)
    desde_info = extract_desde_info(pdf_file_path)
    fechapago1 = date.today().strftime("%m/%d/%Y")
    fecha_pago1_date = datetime.strptime(fechapago1, '%m/%d/%Y')
    fechapago2_date = fecha_pago1_date + relativedelta(months=numero_cuotas)
    fechapago2 = fechapago2_date.strftime('%m/%d/%Y')
    deposito = prima_info * monto_inicial
    if client_info and direccion_info and poliza_info and prima_info and desde_info:
        return {'cliente': client_info, 
                'direccion':direccion_info,
                'poliza':poliza_info,
                'prima': "{0:,.2f}".format(prima_info),
                'desde': desde_info,
                'fechapago1': fechapago1,
                'fechapago2': fechapago2,
                'deposito': "{0:,.2f}".format(deposito),
                'cuotas': numero_cuotas
                 }

def process_pdf_file(pdf_file_path: Path) -> None:
    #Procesa un archivo PDF y genera un archivo DOCX
    print (pdf_file_path)
    pdf_info = extract_pdf_info(pdf_file_path)
    if pdf_info:
        generate_docx(pdf_info)
        file_name = os.path.basename(pdf_file_path)
        file_path = os.path.join(carpeta_procesados, file_name)
        if os.path.exists(file_path):
           os.remove(os.path.join(carpeta_procesados, file_name))
        shutil.move(os.path.join(carpeta_pdf, pdf_file_path), carpeta_procesados)
        
def generate_docx(pdf_info: dict) -> None:
    #Genera un archivo DOCX a partir de la información extraída del PDF
    template = DocxTemplate(plantilla_axeso)
    datos = {
        'direccion': pdf_info['direccion'],
        'cliente': pdf_info['cliente'],
        'prima': pdf_info['prima'],
        'dep': pdf_info['deposito'],
        'cuotas': pdf_info['cuotas'],
        'fechapago1': pdf_info['fechapago1'],
        'fechapago2': pdf_info['fechapago2'],
        'desde': pdf_info['desde'],
        'poliza': pdf_info['poliza']
    }
    template.render(datos)
    nombre_archivo = 'Axeso '+datos['poliza']+'.docx'
    nombre_archivo_pdf = 'Axeso '+datos['poliza']+'.pdf'
    ruta_completa = os.path.join(carpeta_procesados, nombre_archivo)
    ruta_completa_pdf = os.path.join(carpeta_procesados, nombre_archivo_pdf)
    template.save(ruta_completa)
    
    # Convertir el documento a PDF
    docx2pdf.convert(ruta_completa, ruta_completa_pdf)

    # Elimina el archivo Word
    path_docx = os.path.join(carpeta_procesados, nombre_archivo)
    os.remove(path_docx)

class PDFHandler(FileSystemEventHandler):
    def __init__(self, pdf_folder):
        self.pdf_folder = pdf_folder

    def on_created(self, event):
        if event.src_path.endswith('.pdf'):
            process_pdf_file(Path(event.src_path))
    
def main() -> None:
    observer = Observer()
    pdf_handler = PDFHandler(carpeta_pdf)
    observer.schedule(pdf_handler, carpeta_pdf, recursive=True)
    observer.start()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
           