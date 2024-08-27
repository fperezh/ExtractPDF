import openpyxl

# Carga la hoja de Excel
wb = openpyxl.load_workbook('C:/ExtractPDF/CondicionesAxeso.xlsx')
sheet = wb['Condiciones']  # Asigna la hoja que deseas leer

# Define el monto que deseas comparar
monto = 675

# Define el límite superior
limite_superior = 9999999

# Define las celdas que contienen los rangos y factores
rangos = []
factores = []

# Lee los valores de las celdas
for row in range(2, 18):  # desde la fila 2 hasta la 17
    cell = sheet.cell(row=row, column=9)  # columna I es la número 9
    rangos.append(cell.value)

for row in range(2, 18):  # desde la fila 2 hasta la 17
    cell = sheet.cell(row=row, column=11)  # columna K es la número 11
    factores.append(cell.value)

# Implementa la lógica para comparar el monto y obtener el factor correspondiente
for i, rango in enumerate(rangos):
    print(f'Rango: {rango}')
    if monto <= rango:
        factor = factores[i]
        break
    elif monto > limite_superior:
        factor = factores[-1]  # toma el factor del límite superior
        break
else:
    factor = None  # Si no se encuentra un rango que coincida, asigna None

print(f'El factor correspondiente es: {factor}')